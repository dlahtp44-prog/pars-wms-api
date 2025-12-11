import os
import sqlite3
import json
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta # 날짜 계산을 위해 timedelta 추가
import io # 엑셀 파일 처리를 위해 io 추가

from fastapi import FastAPI, Body, Query, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

# 엑셀 업로드용
try:
    import openpyxl
except ImportError:
    openpyxl = None

# --- 설정 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "WMS.db")

app = FastAPI(title="PARS WMS QR BUTTON FINAL")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ----------------------- DB 유틸 -----------------------

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    cur = conn.cursor()

    # 품목 마스터
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
          item_code TEXT PRIMARY KEY,
          item_name TEXT,
          spec      TEXT,
          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    # 로케이션 마스터
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS locations (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          warehouse TEXT NOT NULL,
          location  TEXT NOT NULL,
          UNIQUE(warehouse, location)
        )
        """
    )

    # 재고/거래 이력 테이블
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inventory_tx (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          tx_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          tx_type TEXT NOT NULL, -- IN, OUT, MOVE, OPENING, ROLLBACK
          warehouse TEXT NOT NULL,
          location  TEXT NOT NULL,
          item_code TEXT NOT NULL,
          item_name TEXT,
          lot_no    TEXT NOT NULL,
          qty       REAL NOT NULL,
          remark    TEXT,
          source_tx_id INTEGER, -- ROLLBACK의 경우 원본 TX ID
          UNIQUE(id, tx_type)
        )
        """
    )
    
    # 인덱스 추가 (조회 성능 향상)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_tx_item_wh_loc_lot ON inventory_tx (item_code, warehouse, location, lot_no)")

    conn.commit()
    conn.close()

# DB 초기화는 서버 시작 시 한 번 실행
init_db()


def upsert_item(item_code, item_name):
    conn = get_conn()
    cur = conn.cursor()
    
    # item_name이 빈 문자열("")이면 None으로 변환하여 DB에 NULL로 저장되도록 합니다.
    item_name_to_use = item_name if item_name else None
    
    # 품목 코드가 존재하면 item_name을 업데이트, 없으면 삽입 (항상 업데이트)
    # 엑셀 업로드 시 품명이 비어있으면 NULL로 덮어쓰기 허용
    cur.execute(
        """
        INSERT INTO items (item_code, item_name)
        VALUES (?, ?)
        ON CONFLICT(item_code) DO UPDATE SET item_name=excluded.item_name
        """,
        (item_code, item_name_to_use)
    )
    conn.commit()
    conn.close()


def normalize_lot(lot: Optional[str]) -> str:
    """LOT가 없거나 공백이면 '없음'으로 통일"""
    if lot is None:
        return "없음"
    lot_str = str(lot).strip()
    return lot_str if lot_str else "없음"
    
def get_item_name_by_code(item_code):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT item_name FROM items WHERE item_code = ?", (item_code,))
    row = cur.fetchone()
    conn.close()
    return row["item_name"] if row else ""

# ----------------------- 데이터 모델 -----------------------

class Location(BaseModel):
    warehouse: str
    location: str

class InventoryTx(BaseModel):
    warehouse: str
    location: str
    item_code: str
    item_name: Optional[str] = None
    lot_no: Optional[str] = None
    qty: float = Field(..., ge=0)  # 0 이상 허용, 실제 제약은 로직에서 처리
    remark: Optional[str] = None

class MoveTx(BaseModel):
    warehouse_from: str
    location_from: str
    warehouse_to: str
    location_to: str
    item_code: str
    lot_no: str
    qty: float = Field(..., gt=0)
    remark: Optional[str] = None


# ----------------------- 재고 관리 API -----------------------

@app.post("/api/inventory/in")
def api_inbound(data: InventoryTx):
    conn = get_conn()
    cur = conn.cursor()
    
    # 1. 품명 업데이트 (새로운 품번이면 등록, 기존 품번이면 품명 업데이트 허용)
    upsert_item(data.item_code, data.item_name)
    item_name_final = get_item_name_by_code(data.item_code)
    lot_no = normalize_lot(data.lot_no)
    
    # 2. 거래 기록
    cur.execute(
        """
        INSERT INTO inventory_tx (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
        VALUES ('IN', ?, ?, ?, ?, ?, ?, ?)
        """,
        (data.warehouse, data.location, data.item_code, item_name_final, lot_no, data.qty, data.remark or "입고")
    )
    conn.commit()
    
    # 3. 현재 재고 조회
    cur.execute(
        """
        SELECT SUM(qty) AS current_qty FROM inventory_tx
        WHERE warehouse = ? AND location = ? AND item_code = ? AND lot_no = ?
        """,
        (data.warehouse, data.location, data.item_code, lot_no)
    )
    current_qty = cur.fetchone()["current_qty"]
    conn.close()
    
    return {"result": "OK", "msg": "입고 완료", "현재고": current_qty}


@app.post("/api/inventory/out")
def api_outbound(data: InventoryTx):
    conn = get_conn()
    cur = conn.cursor()
    
    # 1. 현재 재고 확인
    lot_no = normalize_lot(data.lot_no)
    cur.execute(
        """
        SELECT SUM(qty) AS current_qty FROM inventory_tx
        WHERE warehouse = ? AND location = ? AND item_code = ? AND lot_no = ?
        """,
        (data.warehouse, data.location, data.item_code, lot_no)
    )
    current_qty = cur.fetchone()["current_qty"] or 0
    
    if current_qty < data.qty:
        conn.close()
        return {"result": "FAIL", "msg": "재고 부족", "현재고": current_qty}

    # 2. 품명 조회 (출고 시에는 품명을 변경하지 않음)
    item_name_final = get_item_name_by_code(data.item_code)
    
    # 3. 출고 거래 기록 (qty에 -를 붙여 기록)
    cur.execute(
        """
        INSERT INTO inventory_tx (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
        VALUES ('OUT', ?, ?, ?, ?, ?, ?, ?)
        """,
        (data.warehouse, data.location, data.item_code, item_name_final, lot_no, -data.qty, data.remark or "출고")
    )
    conn.commit()
    conn.close()
    
    return {"result": "OK", "msg": "출고 완료", "출고후잔량": current_qty - data.qty}


@app.post("/api/inventory/move")
def api_move(data: MoveTx):
    conn = get_conn()
    cur = conn.cursor()
    
    # 1. 출고지 재고 확인
    cur.execute(
        """
        SELECT SUM(qty) AS current_qty FROM inventory_tx
        WHERE warehouse = ? AND location = ? AND item_code = ? AND lot_no = ?
        """,
        (data.warehouse_from, data.location_from, data.item_code, data.lot_no)
    )
    current_qty = cur.fetchone()["current_qty"] or 0
    
    if current_qty < data.qty:
        conn.close()
        return {"result": "FAIL", "msg": "이동 재고 부족", "현재고": current_qty}

    # 2. 품명 조회
    item_name_final = get_item_name_by_code(data.item_code)
    
    # 3. 출고 거래 기록 (FROM)
    cur.execute(
        """
        INSERT INTO inventory_tx (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
        VALUES ('MOVE', ?, ?, ?, ?, ?, ?, ?)
        """,
        (data.warehouse_from, data.location_from, data.item_code, item_name_final, data.lot_no, -data.qty, f"이동 출고 ({data.warehouse_to}/{data.location_to}) - {data.remark or ''}")
    )
    
    # 4. 입고 거래 기록 (TO)
    cur.execute(
        """
        INSERT INTO inventory_tx (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
        VALUES ('MOVE', ?, ?, ?, ?, ?, ?, ?)
        """,
        (data.warehouse_to, data.location_to, data.item_code, item_name_final, data.lot_no, data.qty, f"이동 입고 ({data.warehouse_from}/{data.location_from}) - {data.remark or ''}")
    )
    conn.commit()
    conn.close()
    
    return {"result": "OK", "msg": "이동 완료"}


@app.post("/api/opening")
def api_opening(data: InventoryTx):
    conn = get_conn()
    cur = conn.cursor()
    
    # 1. 품명 업데이트 (새로운 품번이면 등록, 기존 품번이면 품명 업데이트 허용)
    upsert_item(data.item_code, data.item_name)
    item_name_final = get_item_name_by_code(data.item_code)
    lot_no = normalize_lot(data.lot_no)

    # 2. 거래 기록 (기초 재고는 기존 재고에 합산됨)
    cur.execute(
        """
        INSERT INTO inventory_tx (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
        VALUES ('OPENING', ?, ?, ?, ?, ?, ?, ?)
        """,
        (data.warehouse, data.location, data.item_code, item_name_final, lot_no, data.qty, data.remark or "기초재고")
    )
    conn.commit()
    
    # 3. 현재 재고 조회
    cur.execute(
        """
        SELECT SUM(qty) AS current_qty FROM inventory_tx
        WHERE warehouse = ? AND location = ? AND item_code = ? AND lot_no = ?
        """,
        (data.warehouse, data.location, data.item_code, lot_no)
    )
    current_qty = cur.fetchone()["current_qty"]
    conn.close()
    
    return {"result": "OK", "msg": "기초재고 입력 완료", "현재고": current_qty}


@app.post("/api/rollback")
def api_rollback(tx_id: int = Query(..., description="취소할 거래 ID")):
    conn = get_conn()
    cur = conn.cursor()
    
    # 1. 원본 거래 정보 조회
    cur.execute("SELECT * FROM inventory_tx WHERE id = ?", (tx_id,))
    original_tx = cur.fetchone()
    
    if not original_tx:
        conn.close()
        raise HTTPException(status_code=404, detail="거래 ID를 찾을 수 없습니다.")

    if original_tx["tx_type"] == 'ROLLBACK':
        conn.close()
        raise HTTPException(status_code=400, detail="이미 취소된 거래입니다.")
        
    # 2. 취소 거래 기록 생성
    # 롤백은 원본 거래의 수량에 -1을 곱하여 기록
    rollback_qty = -original_tx["qty"]

    # 3. 재고 부족 확인 (입고, 기초 롤백 시 재고 감소하므로 확인)
    if rollback_qty < 0: 
        cur.execute(
            """
            SELECT SUM(qty) AS current_qty FROM inventory_tx
            WHERE warehouse = ? AND location = ? AND item_code = ? AND lot_no = ?
            """,
            (original_tx["warehouse"], original_tx["location"], original_tx["item_code"], original_tx["lot_no"])
        )
        current_qty = cur.fetchone()["current_qty"] or 0
        
        if current_qty + rollback_qty < 0:
            conn.close()
            raise HTTPException(status_code=400, detail=f"취소 시 재고가 부족해집니다. 현재 재고: {current_qty}")

    # 4. 롤백 거래 기록
    cur.execute(
        """
        INSERT INTO inventory_tx (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark, source_tx_id)
        VALUES ('ROLLBACK', ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            original_tx["warehouse"],
            original_tx["location"],
            original_tx["item_code"],
            original_tx["item_name"],
            original_tx["lot_no"],
            rollback_qty,
            f"TX_ID={tx_id} 롤백",
            tx_id
        )
    )
    conn.commit()
    conn.close()
    
    return {"result": "OK", "msg": f"거래 ID {tx_id} 롤백 완료"}


# ----------------------- Location 관리 API -----------------------

@app.post("/api/location")
def api_create_location(data: Location):
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO locations (warehouse, location) VALUES (?, ?)",
            (data.warehouse, data.location)
        )
        conn.commit()
        return {"id": cur.lastrowid, "msg": "Location 등록 완료"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="이미 존재하는 창고/Location 조합입니다.")
    finally:
        conn.close()

@app.get("/api/location") # URL path changed from /api/location/list to /api/location for consistency (assuming frontend update)
def api_location_list():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM locations ORDER BY warehouse, location")
    rows = cur.fetchall()
    conn.close()
    return [dict(row) for row in rows]

@app.delete("/api/location/{id}")
def api_delete_location(id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM locations WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return {"msg": "Location 삭제 완료"}


# ----------------------- 재고 조회 API -----------------------

@app.get("/api/inventory")
def api_inventory_view(
    item_code: Optional[str] = Query(None),
    lot_no: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None)
):
    conn = get_conn()
    cur = conn.cursor()
    
    where_clauses = []
    params = []
    
    if item_code:
        where_clauses.append("t.item_code LIKE ?")
        params.append(f"%{item_code}%")
    if lot_no:
        where_clauses.append("t.lot_no LIKE ?")
        params.append(f"%{lot_no}%")
    if warehouse:
        where_clauses.append("t.warehouse LIKE ?")
        params.append(f"%{warehouse}%")
    if location:
        where_clauses.append("t.location LIKE ?")
        params.append(f"%{location}%")
        
    # 날짜 필터는 TX 시간 기준
    if date_from:
        where_clauses.append("t.tx_time >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("t.tx_time < ?")
        # 날짜 to는 다음 날 00:00:00으로 설정하여 해당 일자 전체를 포함하도록
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d")
            next_day = dt_to + timedelta(days=1)
            params.append(next_day.strftime("%Y-%m-%d"))
        except ValueError:
            params.append(date_to) # 형식 오류 시 원본 사용

    where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    
    # 마지막 거래 시간 조회를 위한 서브쿼리 추가
    last_tx_subquery = """
        (SELECT MAX(tx_time) FROM inventory_tx WHERE warehouse=t.warehouse AND location=t.location AND item_code=t.item_code AND lot_no=t.lot_no)
    """

    sql = f"""
        SELECT
            t.warehouse,
            t.location,
            t.item_code,
            t.item_name,
            t.lot_no,
            SUM(t.qty) AS current_qty,
            {last_tx_subquery} AS last_tx_time
        FROM inventory_tx t
        {where_sql}
        GROUP BY t.warehouse, t.location, t.item_code, t.lot_no, t.item_name
        HAVING SUM(t.qty) > 0
        ORDER BY t.warehouse, t.location, t.item_code, t.lot_no
    """
    
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    
    # 필터 조회 API를 위해 키 이름을 프론트엔드에 맞춤 (wh->warehouse, loc->location, item->item_code, lot->lot_no)
    # 현재 재고 조회 테이블 렌더링을 위해 키 이름을 유지
    return [dict(row) for row in rows]

# ----------------------- 현재 재고 수량 조회 API (추가) -----------------------

@app.get("/api/inventory/qty")
def api_inventory_qty_view(
    item_code: str = Query(..., description="품번"),
    warehouse: str = Query(..., description="창고"),
    location: str = Query(..., description="Location"),
    lot_no: str = Query(..., description="LOT No.")
):
    conn = get_conn()
    cur = conn.cursor()
    lot_no_normalized = normalize_lot(lot_no)
    
    cur.execute(
        """
        SELECT SUM(qty) AS current_qty FROM inventory_tx
        WHERE warehouse = ? AND location = ? AND item_code = ? AND lot_no = ?
        """,
        (warehouse, location, item_code, lot_no_normalized)
    )
    current_qty = cur.fetchone()["current_qty"] or 0
    conn.close()
    
    return {"current_qty": current_qty}

# ----------------------- 이력 조회 API -----------------------

@app.get("/api/history")
def api_history_view(
    tx_type: Optional[str] = Query(None),
    item_code: Optional[str] = Query(None),
    lot_no: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None)
):
    conn = get_conn()
    cur = conn.cursor()
    
    where_clauses = []
    params = []
    
    if tx_type:
        where_clauses.append("tx_type = ?")
        params.append(tx_type)
    if item_code:
        where_clauses.append("item_code LIKE ?")
        params.append(f"%{item_code}%")
    if lot_no:
        where_clauses.append("lot_no LIKE ?")
        params.append(f"%{lot_no}%")
    if warehouse:
        where_clauses.append("warehouse LIKE ?")
        params.append(f"%{warehouse}%")
    if location:
        where_clauses.append("location LIKE ?")
        params.append(f"%{location}%")
    
    # 날짜 필터 (tx_time)
    if date_from:
        where_clauses.append("tx_time >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("tx_time < ?")
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d")
            next_day = dt_to + timedelta(days=1)
            params.append(next_day.strftime("%Y-%m-%d"))
        except ValueError:
            params.append(date_to)

    where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
    
    sql = f"""
        SELECT
            id,
            strftime('%Y-%m-%d %H:%M:%S', tx_time) as tx_time,
            tx_type,
            warehouse,
            location,
            item_code,
            item_name,
            lot_no,
            qty,
            remark,
            source_tx_id
        FROM inventory_tx
        {where_sql}
        ORDER BY id DESC
    """
    
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    
    return [dict(row) for row in rows]


# ----------------------- 엑셀 업로드 API -----------------------

async def api_upload_location(file: UploadFile):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl 라이브러리가 설치되어 있지 않습니다.")
        
    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 읽기 오류: {e}")
    
    conn = get_conn()
    cur = conn.cursor()
    
    inserted_count = 0
    failed_rows = []
    
    idx = 0
    for row in sheet.iter_rows(values_only=True):
        idx += 1
        if idx == 1:
            continue
            
        vals = [str(v).strip() if v is not None else "" for v in row]
        if not (len(vals) >= 2 and vals[0] and vals[1]):
            if any(vals): # 데이터가 있지만 불완전한 경우
                failed_rows.append({"row": idx, "msg": "필수값(창고/Location) 누락"})
            continue

        wh = vals[0]
        loc = vals[1]
        
        try:
            cur.execute(
                "INSERT INTO locations (warehouse, location) VALUES (?, ?)",
                (wh, loc)
            )
            inserted_count += 1
        except sqlite3.IntegrityError:
            failed_rows.append({"row": idx, "msg": "이미 존재하는 조합"})
        except Exception as e:
            failed_rows.append({"row": idx, "msg": str(e)})

    conn.commit()
    conn.close()
    
    return {"uploaded": idx - 1, "inserted": inserted_count, "failed_rows": failed_rows}

@app.post("/api/upload/{kind}")
async def api_upload_inventory(kind: str, file: UploadFile = File(...)):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl 라이브러리가 설치되어 있지 않습니다.")

    if kind == "location":
        return await api_upload_location(file)
        
    if kind not in ["inbound", "outbound", "move", "opening"]:
        raise HTTPException(status_code=400, detail="유효하지 않은 업로드 유형입니다.")
        
    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 읽기 오류: {e}")
    
    conn = get_conn()
    cur = conn.cursor()
    
    uploaded_count = 0
    failed = []
    
    # 헬퍼 함수 정의
    def process_row(idx, row):
        nonlocal uploaded_count
        
        vals = list(row) if row is not None else []
        
        def val(i):
            return (str(vals[i]).strip() if i < len(vals) and vals[i] is not None else "")

        def safe_float(raw):
            try:
                # 엑셀 셀이 숫자형일 경우 float으로 들어오므로, raw가 None이 아니면 float 변환 시도
                return float(raw)
            except (ValueError, TypeError):
                return None
        
        # 기본 변수 초기화 (모든 케이스에서 사용되는 변수는 여기서 초기화)
        item_code = ""
        item_name = ""
        lot_no = ""

        if kind == "inbound" or kind == "opening":
            # 창고, Location, 품번, 품명, LOT, 수량, 비고
            # vals[5]는 수량으로 가정하고 raw 값을 그대로 사용 (safe_float으로 처리)
            raw_vals = [v for v in vals] 
            wh, loc, item_code, item_name, lot_no_raw, qty_raw, remark = val(0), val(1), val(2), val(3), val(4), raw_vals[5] if len(raw_vals) > 5 else None, val(6)
            qty = safe_float(qty_raw)
            remark = remark or ("엑셀입고" if kind=="inbound" else "엑셀기초")
            lot_no = normalize_lot(lot_no_raw)
            
            # 엑셀 행이 너무 짧거나 필수값 누락/수량 오류
            if not (wh and loc and item_code and lot_no and qty is not None):
                if not any(val(i) for i in range(7)): return # 빈 행 스킵
                failed.append({"row": idx, "msg": "필수값(창고/Loc/품번/LOT/수량) 누락 또는 수량 오류"})
                return

            if qty <= 0 and kind == "inbound": # 입고는 0보다 커야 함
                failed.append({"row": idx, "msg": "입고 수량은 0보다 커야 함"})
                return

            # 품목 마스터 업데이트/삽입 (item_name이 ""이면 NULL로 저장됨)
            upsert_item(item_code, item_name)
            item_name_final = get_item_name_by_code(item_code)
            
            tx_type = "IN" if kind == "inbound" else "OPENING"
            cur.execute(
                """
                INSERT INTO inventory_tx
                (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (tx_type, wh, loc, item_code, item_name_final, lot_no, qty, remark)
            )
            uploaded_count += 1
            
        elif kind == "outbound":
            # 창고, Location, 품번, LOT, 수량, 비고
            raw_vals = [v for v in vals]
            wh, loc, item_code, lot_no_raw, qty_raw, remark = val(0), val(1), val(2), val(3), raw_vals[4] if len(raw_vals) > 4 else None, val(5)
            qty = safe_float(qty_raw)
            remark = remark or "엑셀출고"
            lot_no = normalize_lot(lot_no_raw) # lot_no도 normalize

            if not (wh and loc and item_code and lot_no and qty is not None):
                if not any(val(i) for i in range(6)): return # 빈 행 스킵
                failed.append({"row": idx, "msg": "필수값(창고/Loc/품번/LOT/수량) 누락 또는 수량 오류"})
                return
            
            if qty <= 0:
                failed.append({"row": idx, "msg": "출고 수량은 0보다 커야 함"})
                return
            
            # 엑셀 출고는 별도의 재고 체크 로직을 생략합니다.
            
            item_name_final = get_item_name_by_code(item_code)

            cur.execute(
                """
                INSERT INTO inventory_tx
                (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
                VALUES ('OUT', ?, ?, ?, ?, ?, ?, ?)
                """,
                (wh, loc, item_code, item_name_final, lot_no, -qty, remark)
            )
            uploaded_count += 1
            
        elif kind == "move":
            # 출고창고, 출고Loc, 입고창고, 입고Loc, 품번, LOT, 수량, 비고
            raw_vals = [v for v in vals]
            wh_f, loc_f, wh_t, loc_t, item_code, lot_no_raw, qty_raw, remark = val(0), val(1), val(2), val(3), val(4), val(5), raw_vals[6] if len(raw_vals) > 6 else None, val(7)
            qty = safe_float(qty_raw)
            remark = remark or "엑셀이동"
            lot_no = normalize_lot(lot_no_raw) # lot_no도 normalize
            
            if not (wh_f and loc_f and wh_t and loc_t and item_code and lot_no and qty is not None):
                if not any(val(i) for i in range(8)): return # 빈 행 스킵
                failed.append({"row": idx, "msg": "필수값(출/입고 창고/Loc/품번/LOT/수량) 누락 또는 수량 오류"})
                return

            if qty <= 0:
                failed.append({"row": idx, "msg": "이동 수량은 0보다 커야 함"})
                return

            # 재고 확인 (생략)
            item_name_final = get_item_name_by_code(item_code)

            # 출고 거래 기록 (FROM)
            cur.execute(
                """
                INSERT INTO inventory_tx
                (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
                VALUES ('MOVE', ?, ?, ?, ?, ?, ?, ?)
                """,
                (wh_f, loc_f, item_code, item_name_final, lot_no, -qty, f"엑셀이동 출고 ({wh_t}/{loc_t}) - {remark}")
            )
            # 입고 거래 기록 (TO)
            cur.execute(
                """
                INSERT INTO inventory_tx
                (tx_type, warehouse, location, item_code, item_name, lot_no, qty, remark)
                VALUES ('MOVE', ?, ?, ?, ?, ?, ?, ?)
                """,
                (wh_t, loc_t, item_code, item_name_final, lot_no, qty, f"엑셀이동 입고 ({wh_f}/{loc_f}) - {remark}")
            )
            uploaded_count += 1
        
    
    # 엑셀 시트 행 처리
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1: continue # 헤더 행
        try:
            process_row(idx, row)
        except Exception as e:
            failed.append({"row": idx, "msg": f"내부 오류: {e}"})

    conn.commit()
    conn.close()
    
    return {"uploaded": uploaded_count, "failed": failed}

# ----------------------- 재고 기반 필터 조회 (추가) -----------------------

@app.get("/api/inventory/filters")
def api_inventory_filters(
    item_code: str = Query(..., description="필터링할 품번")
):
    """
    특정 품번(item_code)의 현재 재고를 기준으로
    사용 가능한 창고, 로케이션, 로트 목록을 반환합니다.
    """
    conn = get_conn()
    cur = conn.cursor()

    # 현재 재고가 0보다 큰(SUM(qty) > 0) 모든 고유한 창고/로케이션/LOT 조합을 찾습니다.
    sql = """
      SELECT
        t.warehouse AS wh,
        t.location  AS loc,
        t.lot_no    AS lot
      FROM inventory_tx t
      WHERE t.item_code = ?
      GROUP BY t.warehouse, t.location, t.lot_no
      HAVING SUM(t.qty) > 0
    """
    cur.execute(sql, (item_code,))
    rows = cur.fetchall()
    conn.close()
    
    # 중복 제거를 위해 set을 사용하여 목록 생성
    warehouses = sorted(list(set(r["wh"] for r in rows)))
    locations = sorted(list(set(r["loc"] for r in rows)))
    lots = sorted(list(set(r["lot"] for r in rows)))
    
    # map 구조를 생성하여 프론트엔드에서 종속적인 드롭다운을 구성하기 쉽게 지원
    # { "wh1": { "loc1": ["lotA", "lotB"], "loc2": ["lotA"] }, ... }
    inventory_map = {}
    for r in rows:
        if r['wh'] not in inventory_map:
            inventory_map[r['wh']] = {}
        if r['loc'] not in inventory_map[r['wh']]:
            inventory_map[r['wh']][r['loc']] = []
        inventory_map[r['wh']][r['loc']].append(r['lot'])
        
    # LOT 목록은 중복 제거 후 정렬
    for wh in inventory_map:
        for loc in inventory_map[wh]:
            inventory_map[wh][loc] = sorted(list(set(inventory_map[wh][loc])))


    return {
        "warehouses": warehouses,
        "locations": locations,
        "lots": lots,
        "inventory_map": inventory_map
    }

# ----------------------- 품명 조회 (추가) -----------------------
@app.get("/api/item_name")
def get_item_name_api(item_code: str = Query(..., description="조회할 품번")):
    """품번을 기반으로 품명을 조회합니다."""
    item_name = get_item_name_by_code(item_code)
    return {"item_name": item_name}

